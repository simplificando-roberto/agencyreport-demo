"""AgencyReport v4 - Hardened security."""

import asyncio
import csv
import hashlib
import io
import json
import logging
import os
import secrets
from collections import defaultdict
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from pathlib import Path
from time import time as now
from uuid import UUID

from fastapi import FastAPI, Depends, HTTPException, Request, UploadFile, File, Response
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from jose import jwt, JWTError
from pydantic import BaseModel
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import (
    SECRET_KEY, ALGORITHM, ACCESS_TOKEN_EXPIRE_MINUTES,
    COOKIE_NAME, COOKIE_SECURE, COOKIE_HTTPONLY, COOKIE_SAMESITE,
)
from app.core.database import Base, engine, get_db, SessionLocal
from app.models import Agency, Client, Metric, Alert, Report
from app.services.mock_data import DEMO_CLIENTS, generate_metrics

log = logging.getLogger("agencyreport")
logging.basicConfig(level=logging.INFO)

MAX_UPLOAD_SIZE = 10 * 1024 * 1024  # 10 MB

# ---------------------------------------------------------------------------
# Rate limiter
# ---------------------------------------------------------------------------
_rate_limits: dict[str, list[float]] = defaultdict(list)

def _check_rate_limit(key: str, max_requests: int = 10, window: int = 3600) -> tuple[bool, int]:
    cutoff = now() - window
    _rate_limits[key] = [t for t in _rate_limits[key] if t > cutoff]
    remaining = max_requests - len(_rate_limits[key])
    if remaining <= 0:
        return False, 0
    _rate_limits[key].append(now())
    return True, remaining - 1

# ---------------------------------------------------------------------------
# Password hashing (PBKDF2 with salt - stdlib, no deps)
# ---------------------------------------------------------------------------
def _hash_password(pw: str) -> str:
    salt = os.urandom(16)
    h = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt, 100_000)
    return salt.hex() + ":" + h.hex()

def _verify_password(pw: str, stored: str) -> bool:
    if ":" not in stored:
        # Legacy SHA256 (migration path)
        return hashlib.sha256(pw.encode()).hexdigest() == stored
    salt_hex, hash_hex = stored.split(":", 1)
    h = hashlib.pbkdf2_hmac("sha256", pw.encode(), bytes.fromhex(salt_hex), 100_000)
    return h.hex() == hash_hex

# ---------------------------------------------------------------------------
# Seed
# ---------------------------------------------------------------------------
async def _seed_demo(db: AsyncSession):
    result = await db.execute(select(func.count()).select_from(Agency))
    if result.scalar():
        return
    admin_pw = os.environ.get("ADMIN_PASSWORD", "") or secrets.token_urlsafe(12)
    log.info("Admin account created: admin@agency.test (check ADMIN_PASSWORD env var)")
    agency = Agency(name="Demo Agency", email="admin@agency.test", password_hash=_hash_password(admin_pw), brand_color="#2563eb")
    db.add(agency)
    await db.flush()
    for i, cd in enumerate(DEMO_CLIENTS):
        client = Client(agency_id=agency.id, name=cd["name"], industry=cd["industry"], channels=cd["channels"])
        db.add(client)
        await db.flush()
        for m in generate_metrics(cd["channels"], days=90, client_index=i):
            db.add(Metric(client_id=client.id, **m))
        ch = next(iter(cd["channels"]))
        db.add(Alert(client_id=client.id, channel=ch, metric_name="engagement_rate" if "instagram" in cd["channels"] else "clicks", threshold=2.0, condition="below", triggered=(i == 2)))
    await db.commit()

@asynccontextmanager
async def lifespan(_app: FastAPI):
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    async with SessionLocal() as db:
        await _seed_demo(db)
    yield

app = FastAPI(title="AgencyReport", version="0.4.0", lifespan=lifespan)

# No CORS needed - same origin (frontend served by same server)
# If needed in future, restrict to specific origins

# ---------------------------------------------------------------------------
# Security headers middleware
# ---------------------------------------------------------------------------
@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    return response

# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------
class LoginRequest(BaseModel):
    email: str
    password: str

class LoginResponse(BaseModel):
    agency_name: str
    agency_id: str

class ClientOut(BaseModel):
    id: str
    name: str
    industry: str
    channels: dict
    created_at: datetime

class MetricOut(BaseModel):
    channel: str
    metric_name: str
    value: float
    date: datetime

class DashboardOverview(BaseModel):
    total_clients: int
    active_alerts: int
    total_metrics_today: int
    clients: list[ClientOut]
    recent_alerts: list[dict]

class ReportRequest(BaseModel):
    client_id: str
    period_days: int = 30
    channels: list[str] = []

class ReportOut(BaseModel):
    id: str
    title: str
    period_start: datetime
    period_end: datetime
    ai_summary: str | None
    channels: list
    created_at: datetime

class AlertOut(BaseModel):
    id: str
    client_name: str
    channel: str
    metric_name: str
    threshold: float
    condition: str
    triggered: bool

class ChatRequest(BaseModel):
    message: str
    client_id: str | None = None

class ChatResponse(BaseModel):
    response: str
    status: str
    remaining_requests: int

class TokenInput(BaseModel):
    token: str

# ---------------------------------------------------------------------------
# Auth (cookie-based)
# ---------------------------------------------------------------------------
def _create_token(agency_id: str) -> str:
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    return jwt.encode({"sub": agency_id, "exp": expire}, SECRET_KEY, algorithm=ALGORITHM)

def _set_auth_cookie(response: Response, token: str):
    response.set_cookie(
        COOKIE_NAME, token,
        httponly=COOKIE_HTTPONLY, secure=COOKIE_SECURE,
        samesite=COOKIE_SAMESITE, path="/",
        max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60,
    )

def _clear_auth_cookie(response: Response):
    response.delete_cookie(COOKIE_NAME, path="/")

async def _get_current_agency(request: Request, db: AsyncSession = Depends(get_db)) -> Agency:
    token = request.cookies.get(COOKIE_NAME, "")
    if not token:
        # Fallback to Authorization header (for API clients)
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(401, "Autenticacion requerida")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        agency_id = payload.get("sub")
        if not agency_id:
            raise HTTPException(401, "Token invalido")
    except JWTError:
        raise HTTPException(401, "Token expirado o invalido")
    result = await db.execute(select(Agency).where(Agency.id == UUID(agency_id)))
    agency = result.scalar_one_or_none()
    if not agency:
        raise HTTPException(401, "Agencia no encontrada")
    return agency

# ---------------------------------------------------------------------------
# Auth endpoints
# ---------------------------------------------------------------------------
@app.get("/api/health")
async def health():
    return {"status": "ok"}

@app.post("/api/auth/login", response_model=LoginResponse)
async def login(req: LoginRequest, response: Response, request: Request, db: AsyncSession = Depends(get_db)):
    # Rate limit: 5 attempts per minute per IP
    client_ip = request.client.host if request.client else "unknown"
    allowed, _ = _check_rate_limit(f"login:{client_ip}", max_requests=5, window=60)
    if not allowed:
        raise HTTPException(429, "Demasiados intentos. Espera un minuto.")

    result = await db.execute(select(Agency).where(Agency.email == req.email))
    agency = result.scalar_one_or_none()
    if not agency or not _verify_password(req.password, agency.password_hash):
        raise HTTPException(401, "Credenciales invalidas")

    token = _create_token(str(agency.id))
    _set_auth_cookie(response, token)
    return LoginResponse(agency_name=agency.name, agency_id=str(agency.id))

@app.post("/api/auth/logout")
async def logout(response: Response):
    _clear_auth_cookie(response)
    return {"status": "ok"}

@app.get("/api/auth/me")
async def auth_me(agency: Agency = Depends(_get_current_agency)):
    return {"agency_name": agency.name, "agency_id": str(agency.id)}

# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
@app.get("/api/dashboard/overview", response_model=DashboardOverview)
async def dashboard_overview(db: AsyncSession = Depends(get_db), agency: Agency = Depends(_get_current_agency)):
    clients_r = await db.execute(select(Client).where(Client.agency_id == agency.id))
    clients = clients_r.scalars().all()
    alerts_r = await db.execute(select(func.count()).select_from(Alert).join(Client).where(Client.agency_id == agency.id, Alert.triggered == True))
    recent_r = await db.execute(select(Alert, Client.name).join(Client).where(Client.agency_id == agency.id, Alert.triggered == True).order_by(Alert.created_at.desc()).limit(3))
    recent_alerts = [{"client": cn, "channel": a.channel, "metric": a.metric_name, "threshold": a.threshold} for a, cn in recent_r.all()]
    return DashboardOverview(
        total_clients=len(clients), active_alerts=alerts_r.scalar() or 0, total_metrics_today=len(clients) * 5,
        clients=[ClientOut(id=str(c.id), name=c.name, industry=c.industry, channels=c.channels, created_at=c.created_at) for c in clients],
        recent_alerts=recent_alerts,
    )

# ---------------------------------------------------------------------------
# Clients & Metrics
# ---------------------------------------------------------------------------
@app.get("/api/clients", response_model=list[ClientOut])
async def list_clients(db: AsyncSession = Depends(get_db), agency: Agency = Depends(_get_current_agency)):
    result = await db.execute(select(Client).where(Client.agency_id == agency.id))
    return [ClientOut(id=str(c.id), name=c.name, industry=c.industry, channels=c.channels, created_at=c.created_at) for c in result.scalars().all()]

@app.get("/api/clients/{client_id}/metrics", response_model=list[MetricOut])
async def client_metrics(client_id: str, period: int = 30, channel: str | None = None, db: AsyncSession = Depends(get_db), _a: Agency = Depends(_get_current_agency)):
    cutoff = datetime.utcnow() - timedelta(days=min(period, 365))
    query = select(Metric).where(Metric.client_id == UUID(client_id), Metric.date >= cutoff).order_by(Metric.date)
    if channel:
        query = query.where(Metric.channel == channel)
    result = await db.execute(query)
    return [MetricOut(channel=m.channel, metric_name=m.metric_name, value=m.value, date=m.date) for m in result.scalars().all()]

@app.get("/api/clients/{client_id}/metrics/csv")
async def export_metrics_csv(client_id: str, period: int = 90, db: AsyncSession = Depends(get_db), _a: Agency = Depends(_get_current_agency)):
    cutoff = datetime.utcnow() - timedelta(days=min(period, 365))
    result = await db.execute(select(Metric).where(Metric.client_id == UUID(client_id), Metric.date >= cutoff).order_by(Metric.date))
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["fecha", "canal", "metrica", "valor"])
    for m in result.scalars().all():
        writer.writerow([m.date.strftime("%Y-%m-%d"), m.channel, m.metric_name, m.value])
    buf.seek(0)
    return StreamingResponse(io.BytesIO(buf.getvalue().encode("utf-8-sig")), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=metricas_{client_id[:8]}.csv"})

# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------
@app.post("/api/reports/generate", response_model=ReportOut)
async def generate_report(req: ReportRequest, db: AsyncSession = Depends(get_db), agency: Agency = Depends(_get_current_agency)):
    n = datetime.utcnow()
    client_r = await db.execute(select(Client).where(Client.id == UUID(req.client_id)))
    client = client_r.scalar_one_or_none()
    if not client:
        raise HTTPException(404, "Cliente no encontrado")
    channels = req.channels or list(client.channels.keys())
    ai_summary = (f"Resumen ejecutivo de {client.name} ({client.industry}) - Ultimos {req.period_days} dias:\n\n"
                  f"Los canales {', '.join(channels)} muestran una tendencia positiva con crecimiento sostenido.")
    report = Report(client_id=client.id, agency_id=agency.id, title=f"Reporte {client.name} - {n.strftime('%B %Y')}", period_start=n - timedelta(days=req.period_days), period_end=n, ai_summary=ai_summary, channels=channels)
    db.add(report)
    await db.commit()
    await db.refresh(report)
    return ReportOut(id=str(report.id), title=report.title, period_start=report.period_start, period_end=report.period_end, ai_summary=report.ai_summary, channels=report.channels, created_at=report.created_at)

@app.get("/api/reports", response_model=list[ReportOut])
async def list_reports(skip: int = 0, limit: int = 50, db: AsyncSession = Depends(get_db), agency: Agency = Depends(_get_current_agency)):
    result = await db.execute(select(Report).where(Report.agency_id == agency.id).order_by(Report.created_at.desc()).offset(skip).limit(min(limit, 100)))
    return [ReportOut(id=str(r.id), title=r.title, period_start=r.period_start, period_end=r.period_end, ai_summary=r.ai_summary, channels=r.channels, created_at=r.created_at) for r in result.scalars().all()]

@app.get("/api/reports/{report_id}/excel")
async def download_report_excel(report_id: str, db: AsyncSession = Depends(get_db), _a: Agency = Depends(_get_current_agency)):
    from openpyxl import Workbook
    report_r = await db.execute(select(Report).where(Report.id == UUID(report_id)))
    report = report_r.scalar_one_or_none()
    if not report:
        raise HTTPException(404, "Reporte no encontrado")
    metrics_r = await db.execute(select(Metric).where(Metric.client_id == report.client_id, Metric.date >= report.period_start, Metric.date <= report.period_end).order_by(Metric.date))
    wb = Workbook()
    ws = wb.active
    if ws:
        ws.title = "Metricas"
        ws.append(["Fecha", "Canal", "Metrica", "Valor"])
        for m in metrics_r.scalars().all():
            ws.append([m.date.strftime("%Y-%m-%d"), m.channel, m.metric_name, m.value])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=reporte_{report_id[:8]}.xlsx"})

@app.post("/api/data/upload")
async def upload_data(client_id: str, file: UploadFile = File(...), db: AsyncSession = Depends(get_db), _a: Agency = Depends(_get_current_agency)):
    content = await file.read()
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(413, "Archivo demasiado grande (max 10MB)")
    rows_imported = 0
    if file.filename and file.filename.endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        if not ws:
            raise HTTPException(400, "Excel vacio")
        headers = [str(c.value or "").lower().strip() for c in ws[1]]
        if "fecha" not in headers or "valor" not in headers:
            raise HTTPException(400, "Columnas requeridas: fecha, valor")
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            if not data.get("fecha") or not data.get("valor"):
                continue
            try:
                date_val = data["fecha"] if isinstance(data["fecha"], datetime) else datetime.strptime(str(data["fecha"])[:10], "%Y-%m-%d")
                val = float(data["valor"])
            except (ValueError, TypeError):
                continue
            db.add(Metric(client_id=UUID(client_id), channel=str(data.get("canal", "manual"))[:50], metric_name=str(data.get("metrica", "valor"))[:100], value=val, date=date_val))
            rows_imported += 1
    else:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            date_str = (row.get("fecha", "") or row.get("date", ""))[:10]
            value_str = row.get("valor", "") or row.get("value", "")
            if not date_str or not value_str:
                continue
            try:
                val = float(value_str)
                date_val = datetime.strptime(date_str, "%Y-%m-%d")
            except (ValueError, TypeError):
                continue
            db.add(Metric(client_id=UUID(client_id), channel=str(row.get("canal", "") or row.get("channel", "manual"))[:50], metric_name=str(row.get("metrica", "") or row.get("metric", "valor"))[:100], value=val, date=date_val))
            rows_imported += 1
    await db.commit()
    return {"status": "ok", "rows_imported": rows_imported}

# ---------------------------------------------------------------------------
# Alerts
# ---------------------------------------------------------------------------
@app.get("/api/alerts", response_model=list[AlertOut])
async def list_alerts(db: AsyncSession = Depends(get_db), agency: Agency = Depends(_get_current_agency)):
    result = await db.execute(select(Alert, Client.name).join(Client).where(Client.agency_id == agency.id))
    return [AlertOut(id=str(a.id), client_name=cn, channel=a.channel, metric_name=a.metric_name, threshold=a.threshold, condition=a.condition, triggered=a.triggered) for a, cn in result.all()]

# ---------------------------------------------------------------------------
# AI Setup
# ---------------------------------------------------------------------------
_AI_CONFIG_PATH = Path("/app/config/ai_config.json") if Path("/app/config").is_dir() else Path("/tmp/ai_config.json")

def _load_ai_config() -> dict:
    if _AI_CONFIG_PATH.exists():
        return json.loads(_AI_CONFIG_PATH.read_text())
    return {"default_provider": "claude"}

def _save_ai_config(cfg: dict):
    _AI_CONFIG_PATH.write_text(json.dumps(cfg, indent=2))

async def _check_cli_status(cli: str) -> dict:
    result = {"installed": False, "authenticated": False, "version": ""}
    try:
        proc = await asyncio.create_subprocess_exec(cli, "--version", stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE)
        stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=5)
        if proc.returncode == 0:
            result["installed"] = True
            result["version"] = stdout.decode().strip()
    except (FileNotFoundError, asyncio.TimeoutError):
        return result
    try:
        if cli == "claude":
            proc = await asyncio.create_subprocess_exec("claude", "auth", "status", "--json", stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE)
            stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=5)
            data = json.loads(stdout.decode())
            result["authenticated"] = data.get("loggedIn", False)
        elif cli == "codex":
            proc = await asyncio.create_subprocess_exec("codex", "login", "status", stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE)
            stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=5)
            result["authenticated"] = proc.returncode == 0 and "logged in" in stdout.decode().lower()
    except Exception:
        pass
    return result

@app.get("/api/ai/setup/status")
async def ai_setup_status(_a: Agency = Depends(_get_current_agency)):
    claude_status = await _check_cli_status("claude")
    codex_status = await _check_cli_status("codex")
    config = _load_ai_config()
    return {
        "default_provider": config.get("default_provider", "claude"),
        "providers": {
            "claude": {**claude_status, "name": "Claude Code"},
            "codex": {**codex_status, "name": "Codex (OpenAI)"},
        },
    }

@app.post("/api/ai/setup/login")
async def ai_setup_login(req: TokenInput, provider: str = "claude", _a: Agency = Depends(_get_current_agency)):
    token_val = req.token.strip()
    if not token_val or len(token_val) < 10 or len(token_val) > 500:
        raise HTTPException(400, "Token invalido")
    try:
        if provider == "claude":
            proc = await asyncio.create_subprocess_exec("claude", "setup-token", stdin=asyncio.subprocess.PIPE, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE)
        elif provider == "codex":
            proc = await asyncio.create_subprocess_exec("codex", "login", "--with-api-key", stdin=asyncio.subprocess.PIPE, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE)
        else:
            raise HTTPException(400, "Proveedor desconocido")
        await asyncio.wait_for(proc.communicate(input=f"{token_val}\n".encode()), timeout=15)
        await asyncio.sleep(1)
        status = await _check_cli_status(provider)
        return {"provider": provider, "authenticated": status["authenticated"], "message": "Conectado!" if status["authenticated"] else "Token no valido"}
    except FileNotFoundError:
        raise HTTPException(400, f"{provider} no instalado")

@app.post("/api/ai/setup/verify")
async def ai_setup_verify(provider: str = "claude", _a: Agency = Depends(_get_current_agency)):
    status = await _check_cli_status(provider)
    return {"provider": provider, "authenticated": status["authenticated"]}

@app.post("/api/ai/setup/default")
async def ai_setup_set_default(provider: str = "claude", _a: Agency = Depends(_get_current_agency)):
    config = _load_ai_config()
    config["default_provider"] = provider
    _save_ai_config(config)
    return {"default_provider": provider}

@app.post("/api/ai/setup/logout")
async def ai_setup_logout(provider: str = "claude", _a: Agency = Depends(_get_current_agency)):
    try:
        if provider == "claude":
            proc = await asyncio.create_subprocess_exec("claude", "auth", "logout", stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE)
            await asyncio.wait_for(proc.communicate(), timeout=10)
    except Exception:
        pass
    return {"status": "logged_out"}

# ---------------------------------------------------------------------------
# AI Chat
# ---------------------------------------------------------------------------
@app.get("/api/ai/suggestions")
async def ai_suggestions(client_id: str | None = None, _a: Agency = Depends(_get_current_agency)):
    base = ["Como ha sido el rendimiento general este mes?", "Que canal tiene mejor ROI?", "Genera un resumen ejecutivo", "Que mejoras recomiendas?"]
    if client_id:
        return {"suggestions": base + ["Compara el ultimo mes con el anterior"]}
    return {"suggestions": base}

async def _get_clients_summary(db: AsyncSession, agency_id) -> list[dict]:
    clients_r = await db.execute(select(Client).where(Client.agency_id == agency_id))
    result = []
    cutoff = datetime.utcnow() - timedelta(days=30)
    for c in clients_r.scalars().all():
        metrics_r = await db.execute(select(Metric).where(Metric.client_id == c.id, Metric.date >= cutoff).order_by(Metric.date.desc()).limit(20))
        summary = {}
        for m in metrics_r.scalars().all():
            key = f"{m.channel}/{m.metric_name}"
            if key not in summary:
                summary[key] = {"channel": m.channel, "metric": m.metric_name, "value": m.value}
        result.append({"name": c.name, "industry": c.industry, "channels": list(c.channels.keys()), "metrics": list(summary.values())[:10]})
    return result

@app.post("/api/ai/chat", response_model=ChatResponse)
async def ai_chat(req: ChatRequest, db: AsyncSession = Depends(get_db), agency: Agency = Depends(_get_current_agency)):
    allowed, remaining = _check_rate_limit(str(agency.id), max_requests=10, window=3600)
    if not allowed:
        raise HTTPException(429, "Limite alcanzado (10/hora)")

    config = _load_ai_config()
    provider = config.get("default_provider", "claude")
    if provider not in ("claude", "codex"):
        provider = "claude"

    # Build context
    if req.client_id:
        client_r = await db.execute(select(Client).where(Client.id == UUID(req.client_id)))
        client = client_r.scalar_one_or_none()
        if client:
            cutoff = datetime.utcnow() - timedelta(days=30)
            metrics_r = await db.execute(select(Metric).where(Metric.client_id == UUID(req.client_id), Metric.date >= cutoff).order_by(Metric.date.desc()).limit(20))
            context = f"Cliente: {client.name} ({client.industry})\nMetricas:\n"
            for m in metrics_r.scalars().all():
                context += f"  {m.date.strftime('%d/%m')}: {m.channel}/{m.metric_name} = {m.value:.1f}\n"
        else:
            context = "Cliente no encontrado"
    else:
        clients_summary = await _get_clients_summary(db, agency.id)
        context = json.dumps(clients_summary, indent=2, ensure_ascii=False)

    system_prompt = f"Eres el asistente de AgencyReport. Responde en espanol. Datos:\n{context}"

    try:
        if provider == "claude":
            cmd = ["claude", "--print", "-p", f"{system_prompt}\n\nUsuario: {req.message}"]
        else:
            cmd = ["codex", "exec", req.message]
        proc = await asyncio.create_subprocess_exec(*cmd, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE, cwd="/app")
        stdout_data, _ = await asyncio.wait_for(proc.communicate(), timeout=120)
        text = stdout_data.decode("utf-8", errors="replace").strip()
        if not text:
            text = "Sin respuesta. Verifica que el proveedor este autenticado en Configuracion."
        return ChatResponse(response=text, status="ok", remaining_requests=remaining)
    except asyncio.TimeoutError:
        return ChatResponse(response="Timeout.", status="timeout", remaining_requests=remaining)
    except FileNotFoundError:
        return ChatResponse(response=f"{provider} no disponible. Ve a Configuracion.", status="unavailable", remaining_requests=remaining)

# ---------------------------------------------------------------------------
# Static files - MUST be last
# ---------------------------------------------------------------------------
_static_dir = Path("/app/static")
if _static_dir.is_dir():
    app.mount("/", StaticFiles(directory=str(_static_dir), html=True), name="frontend")
